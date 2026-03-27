import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import locale

# Tenta configurar para português para as datas, se falhar usa mapeamento manual
try:
    locale.setlocale(locale.LC_TIME, "pt_BR.utf8")
except:
    pass

meses_pt = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março",
    "April": "Abril", "May": "Maio", "June": "Junho",
    "July": "Julho", "August": "Agosto", "September": "Setembro",
    "October": "Outubro", "November": "Novembro", "December": "Dezembro"
}

# Configuração da página
st.set_page_config(page_title="PPC - Sistema Fiscal", page_icon="📊", layout="wide")

# --- INICIALIZAÇÃO DO ESTADO GLOBAL ---
if 'dados_processados' not in st.session_state:
    st.session_state.dados_processados = None

# --- FUNÇÃO ORIGINAL DA MEMÓRIA DE CÁLCULO (ESTRUTURA PRESERVADA) ---
def aplicar_estilo_ppc(writer, df_filtrado, colunas_mapeadas, nome_aba, titulo_imposto, razao, cnpj, comp):
    ws = writer.book.create_sheet(nome_aba)
    writer.sheets[nome_aba] = ws
    ws.sheet_view.showGridLines = False 
    ws.column_dimensions['A'].width = 3

    fill_azul = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    font_branca = Font(color='FFFFFF', bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')

    textos_cabecalho = [
        f'RAZÃO SOCIAL: {razao}', 
        f'CNPJ: {cnpj}', 
        f'{titulo_imposto} - COMPETÊNCIA {comp}'
    ]
    
    ultima_col_idx = len(colunas_mapeadas) + 1
    
    for row_num, texto in enumerate(textos_cabecalho, 2):
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=ultima_col_idx)
        cell = ws.cell(row=row_num, column=2)
        cell.value = texto
        cell.alignment = align_center
        cell.font = Font(bold=True, size=12)

    for col_num, header in enumerate(colunas_mapeadas.values(), 2):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = fill_azul
        cell.font = font_branca
        cell.alignment = align_center
        cell.border = thin_border

    total_valor_imposto = 0 # Para alimentar o consolidado

    if df_filtrado.empty:
        row_msg = 7
        ws.merge_cells(start_row=row_msg, start_column=2, end_row=row_msg, end_column=ultima_col_idx)
        cell_msg = ws.cell(row=row_msg, column=2, value="SEM MOVIMENTO")
        cell_msg.font = Font(bold=True)
        cell_msg.alignment = align_center
        for col_idx in range(2, ultima_col_idx + 1):
            ws.cell(row=row_msg, column=col_idx).border = thin_border
    else:
        # TRATAMENTO DO CÓDIGO DE SERVIÇO (Ponto em vez de vírgula)
        for col_orig, col_dest in colunas_mapeadas.items():
            if col_dest == 'Cód. Serviço':
                df_filtrado[col_orig] = df_filtrado[col_orig].astype(str).str.replace(',', '.')

        dados_finais = df_filtrado[list(colunas_mapeadas.keys())].rename(columns=colunas_mapeadas)
        moeda_cols = ['Vlr Contábil', 'Base IRRF', 'Valor IRRF', 'Base CSR', 'Total PCC', 'ISS', 'Valor INSS', 'Base INSS', 'Base ISS']
        
        for r_idx, row in enumerate(dados_finais.values, 7):
            for c_idx, value in enumerate(row, 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = align_center 
                
                header_text = list(colunas_mapeadas.values())[c_idx-2]
                if header_text in moeda_cols:
                    cell.number_format = 'R$ #,##0.00'
                elif 'Data' in header_text:
                    cell.number_format = 'dd/mm/yyyy'
                elif 'Aliq' in header_text or '%' in header_text:
                    cell.number_format = '0.00%'

        last_row = 6 + len(dados_finais)
        row_total = last_row + 1
        
        # BORDAS E FORMATAÇÃO DO TOTAL (CORRIGIDO)
        ws.merge_cells(start_row=row_total, start_column=2, end_row=row_total, end_column=6)
        cell_total_label = ws.cell(row=row_total, column=2, value="TOTAL")
        cell_total_label.font = Font(bold=True)
        cell_total_label.alignment = Alignment(horizontal='right', vertical='center')
        
        for col_idx in range(2, ultima_col_idx + 1):
            ws.cell(row=row_total, column=col_idx).border = thin_border

        for col_idx in range(7, ultima_col_idx + 1):
            header_text = list(colunas_mapeadas.values())[col_idx-2]
            if header_text in moeda_cols:
                col_letter = get_column_letter(col_idx)
                cell_sum = ws.cell(row=row_total, column=col_idx)
                # Soma via fórmula Excel para a planilha, mas calculamos aqui para o consolidado
                valor_soma = dados_finais[header_text].sum()
                cell_sum.value = f"=SUM({col_letter}7:{col_letter}{last_row})"
                cell_sum.font = Font(bold=True)
                cell_sum.number_format = 'R$ #,##0.00'
                
                # Captura o valor se for a coluna de imposto final
                if header_text in ['Valor IRRF', 'Total PCC', 'Valor INSS', 'ISS']:
                    total_valor_imposto += valor_soma

    # Ajuste de largura automático
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        if column_letter == 'A': continue
        for cell in col:
            if cell.row >= 6 and cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_length: max_length = len(val_str)
        ws.column_dimensions[column_letter].width = max_length + 4

    return total_valor_imposto

# --- 1. INTERFACE MEMÓRIA DE CÁLCULO ---
def gerador_memoria_calculo():
    st.title("📊 1. Gerador de Memória de Cálculo")
    arquivo_upload = st.file_uploader("Arraste a planilha do sistema Unecont aqui:", type=["xlsx"], key="up_memoria")

    if arquivo_upload:
        try:
            df = pd.read_excel(arquivo_upload)
            df['ISS_TOTAL'] = df['ISS Dentro do Município'].fillna(0) + df['ISS Fora do Município'].fillna(0)
            df['BASE_ISS_TOTAL'] = df['Base de Cálculo ISS'].fillna(0)
            df['ALIQ_ISS_TOTAL'] = df['% ISS Dentro do Município'].fillna(0) + df['% ISS Fora do Município'].fillna(0)

            razao = df['Empresa'].iloc[0]
            cnpj = df['Cnpj Empresa'].iloc[0]
            data_dt = pd.to_datetime(df['Data Competência'].iloc[0])
            
            # Correção de Idioma para o Período
            mes_eng = data_dt.strftime('%B')
            periodo_pt = f"{meses_pt.get(mes_eng, mes_eng)} de {data_dt.year}"
            comp_file = data_dt.strftime('%m.%Y')
            comp_titulo = data_dt.strftime('%m/%Y')

            output = BytesIO()
            resumo = {}

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                m_base = {'Emissão NFe': 'Data Emissão', 'Número NFe': 'Nota Fiscal', 'Serviço Federal': 'Cód. Serviço', 'Prestador': 'Prestador', 'Cnpj/Cpf Prestador': 'CNPJ', 'Valor NFe': 'Vlr Contábil'}
                
                # Mapeamentos completos (com alíquotas preservadas)
                resumo['1708'] = aplicar_estilo_ppc(writer, df[df['DARF IRRF'] == 1708], {**m_base, 'Base de Cálculo ISS': 'Base IRRF', '% IRRF': 'Aliq. IRRF', 'Valor IRRF': 'Valor IRRF'}, 'IRRF 1708', 'IRRF 1708', razao, cnpj, comp_titulo)
                resumo['5952'] = aplicar_estilo_ppc(writer, df[df['DARF CSRF'] == 5952], {**m_base, 'Base de Cálculo ISS': 'Base CSR', '% CSRF': 'Aliq. CSRF', 'Valor CSRF': 'Total PCC'}, 'CSRF', 'CSRF', razao, cnpj, comp_titulo)
                resumo['8045'] = aplicar_estilo_ppc(writer, df[df['DARF IRRF'] == 8045], {**m_base, 'Base de Cálculo ISS': 'Base IRRF', '% IRRF': 'Aliq. IRRF', 'Valor IRRF': 'Valor IRRF'}, 'IRRF 8045', 'IRRF 8045', razao, cnpj, comp_titulo)
                resumo['3208'] = aplicar_estilo_ppc(writer, df[df['DARF IRRF'] == 3208], {**m_base, 'Base de Cálculo ISS': 'Base IRRF', '% IRRF': 'Aliq. IRRF', 'Valor IRRF': 'Valor IRRF'}, 'IRRF 3208', 'IRRF 3208', razao, cnpj, comp_titulo)
                resumo['ISS'] = aplicar_estilo_ppc(writer, df[df['ISS_TOTAL'] > 0], {**m_base, 'BASE_ISS_TOTAL': 'Base ISS', 'ALIQ_ISS_TOTAL': 'Aliq. ISS', 'ISS_TOTAL': 'ISS'}, 'ISS', 'ISS', razao, cnpj, comp_titulo)
                resumo['INSS_RET'] = aplicar_estilo_ppc(writer, df[df['Valor INSS'] > 0], {**m_base, 'Base de Cálculo INSS': 'Base INSS', '% INSS': 'Aliq. INSS', 'Valor INSS': 'Valor INSS'}, 'INSS', 'INSS', razao, cnpj, comp_titulo)

            st.session_state.dados_processados = {
                'razao': razao, 'cnpj': cnpj, 'periodo': periodo_pt, 'periodo_file': comp_file, 'valores': resumo
            }

            st.success(f"✅ Memória de Cálculo de {razao} processada!")
            st.download_button(label="📥 Baixar Memória de Cálculo", data=output.getvalue(), file_name=f"{razao} - Memoria de Calculo Retidos {comp_file}.xlsx")
        
        except Exception as e:
            st.error(f"Erro ao processar: {e}")

# --- 2. INTERFACE RELATÓRIO CONSOLIDADO ---
def gerador_relatorio_consolidado():
    st.title("📄 2. Relatório Mensal Consolidado")
    
    if st.session_state.dados_processados is None:
        st.warning("⚠️ Processe a 'Memória de Cálculo' primeiro.")
        return

    dados = st.session_state.dados_processados
    st.info(f"Gerando para: **{dados['razao']}** | Competência: **{dados['periodo']}**")

    if st.button("Gerar Relatório Consolidado"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Consolidado"
        ws.sheet_view.showGridLines = False

        # ESTILOS
        azul_ppc = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        cinza_claro = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        laranja_suave = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        font_branca_bold = Font(color='FFFFFF', bold=True)
        font_preta_bold = Font(color='000000', bold=True)
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col, w in zip(['A','B','C','D','E','F','G'], [2, 12, 15, 15, 25, 60, 40]):
            ws.column_dimensions[col].width = w

        ws.merge_cells('B2:G2')
        ws['B2'].value = "DCTFWeb - Relatório Mensal de Impostos Federais Consolidados"
        ws['B2'].font, ws['B2'].fill, ws['B2'].alignment = font_branca_bold, azul_ppc, Alignment(horizontal='center', vertical='center')

        id_info = [
            ("Razão social", dados['razao']),
            ("CNPJ", dados['cnpj']),
            ("Período de apuração", dados['periodo']),
            ("Responsável preenchimento", "MARCOS PAULO SANTOS DE OLIVEIRA")
        ]

        for i, (label, valor) in enumerate(id_info, 4):
            ws.merge_cells(f'B{i}:D{i}'); ws.merge_cells(f'E{i}:G{i}')
            ws[f'B{i}'].value, ws[f'B{i}'].fill, ws[f'B{i}'].font, ws[f'B{i}'].alignment = label, cinza_claro, font_preta_bold, Alignment(horizontal='center')
            ws[f'E{i}'].value, ws[f'E{i}'].fill, ws[f'E{i}'].alignment = valor, cinza_claro, Alignment(horizontal='left', indent=1)
            if label == "Período de apuração": ws[f'E{i}'].font = font_preta_bold

        headers = ["Tipo", "Código Retenção", "Valor Retenção", "Descrição do Código da Receita", "", "Observações"]
        for col, text in zip(range(2, 8), headers):
            cell = ws.cell(row=9, column=col, value=text)
            cell.font, cell.fill, cell.border, cell.alignment = font_branca_bold, azul_ppc, borda_fina, Alignment(horizontal='center')
        ws.merge_cells('E9:F9')

        v = dados['valores']
        impostos = [
            ("INSS", "Folha", 0, "Informação transmitida via eSocial", "Evidência enviada pelo RH"),
            ("IRRF", "0588", 0, "IRRF - Rendimento do Trabalho sem Vínculo Empregatício", ""),
            ("IRRF", "0561", 0, "IRRF - Rendimento do Trabalho Assalariado", ""),
            ("INSS", "1162", v.get('INSS_RET', 0), "Informação transmitida via EFD REINF - Retenção NFSe", "Memória de cálculo do fiscal"),
            ("IRRF", "1708", v.get('1708', 0), "IRRF - Remuneração Serviços Prestados por PJ", ""),
            ("IRRF", "8045", v.get('8045', 0), "IRRF - Outros Rendimentos", ""),
            ("IRRF", "3208", v.get('3208', 0), "IRRF - Aluguéis e Royalties Pagos a PF", ""),
            ("CSRF", "5952", v.get('5952', 0), "Retenção de Contribuições (CSLL, Cofins e PIS)", ""),
            ("PIS", "8109", 0, "PIS - FATURAMENTO - PJ EM GERAL", ""),
            ("COFINS", "2172", 0, "COFINS - FATURAMENTO/PJ EM GERAL", ""),
            ("IRPJ", "2089", 0, "IRPJ - LUCRO PRESUMIDO", ""),
            ("CSLL", "2372", 0, "CSLL - LUCRO PRESUMIDO OU ARBITRADO", "")
        ]

        total_darf, row_idx = 0, 10
        for tipo, cod, valor, desc, obs in impostos:
            total_darf += valor
            c_tipo = ws.cell(row=row_idx, column=2, value=tipo)
            c_tipo.fill, c_tipo.font = laranja_suave, font_preta_bold
            ws.cell(row=row_idx, column=3, value=cod)
            ws.cell(row=row_idx, column=4, value=valor).number_format = 'R$ #,##0.00'
            ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=5, value=desc)
            ws.cell(row=row_idx, column=7, value=obs)
            for c in range(2, 8):
                ws.cell(row=row_idx, column=c).border = borda_fina
                ws.cell(row=row_idx, column=c).alignment = Alignment(horizontal='center')
            row_idx += 1

        ws.merge_cells(f'B{row_idx}:C{row_idx}')
        ws[f'B{row_idx}'].value, ws[f'B{row_idx}'].fill, ws[f'B{row_idx}'].font, ws[f'B{row_idx}'].alignment = "Valor Total DARF", azul_ppc, font_branca_bold, Alignment(horizontal='center')
        c_total = ws.cell(row=row_idx, column=4, value=total_darf)
        c_total.font, c_total.number_format, c_total.border, c_total.alignment = Font(bold=True, color='FF0000'), 'R$ #,##0.00', borda_fina, Alignment(horizontal='center')
        ws.merge_cells(f'E{row_idx}:G{row_idx}')
        ws[f'E{row_idx}'].fill = cinza_claro
        for c in [5, 6, 7]: ws.cell(row=row_idx, column=c).border = borda_fina

        output = BytesIO()
        wb.save(output)
        st.download_button(label="📥 Baixar Relatório Consolidado", data=output.getvalue(), file_name=f"{dados['razao']} - Relatorio Mensal de Impostos Federais Consolidados - {dados['periodo_file']}.xlsx")

# --- LÓGICA DE NAVEGAÇÃO ---
st.sidebar.title("Menu PPC")
opcao = st.sidebar.radio("Escolha a etapa:", ["Memória de Cálculo", "Relatório Consolidado"])

if opcao == "Memória de Cálculo":
    gerador_memoria_calculo()
else:
    gerador_relatorio_consolidado()
